from selenium import webdriver
import time
import openpyxl

PATH = ''  # Download edge driver for your version and replace its path here
driver = webdriver.Edge(PATH)
wb = openpyxl.load_workbook("")  # Replace the path of xlsx file, where you want to store the data
sh1 = wb.active

# first login into moodle
driver.get("https://moodle.iitd.ac.in/")
time.sleep(20)

loc = 2

# replace these number wih the range of indexes you want to scrape data from

for index in range(123, 456):
    try:
        driver.get(f"https://a.impartus.com/ilc/#/profile/{index}")
        name = driver.find_element_by_class_name("md-subhead").text
        karma = driver.find_element_by_class_name("md-body-1").text
        print(name)
        print(karma)
        sh1.cell(column=1, row=loc, value=name)
        sh1.cell(column=2, row=loc, value=karma)
        loc += 1
    except:
        pass
    time.sleep(1)

driver.close()
